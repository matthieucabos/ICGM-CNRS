import os
import sys
import pyexcel as p
import openpyxl as op
import re

__author__="CABOS Matthieu"
__date__=23/12/2021

Vlans={
501:"IDRAC",
510:"DPT1",
511:"DPT2",
512:"DPT3",
513:"DPT4",
514:"DPT5",
515:"INSTRU-ON",
516:"INSTRU-OFF",
518:"IMPRIM",
519:"GUEST",
524:"SGAF",
525:"SSI",
526:"ExpProtect",
528:"IDRAC",
529:"Did",
530:"PT"
}

def Del_Duplicate(liste):
	verif=liste[:]
	for item in liste:
		verif.remove(item)
		if item in verif:
			liste.remove(item)
	return liste

def Get_Users_Info():

	# Building DHCP dictionnary and get infos since the given IP adresses list as parameter

	# Variable Initialisation

	tmp_dict={}
	Users=[]
	Users_dict={}
	DHCP_Dict={}
	Content=""
	tmp=""
	socket=""
	count=0	

	# Regular Expressions Definition

	regex_MAC=r'([0-9A-Fa-f]{2}\:){5}[0-9A-Fa-f]{2}'
	regex_IP=r'fixed.*'
	regex_raw_ip=r'([0-9]+\.){3}[0-9]+'
	regex_hostname=r'\"[A-Za-z0-9-_]+\"'
	regex_cisco=r'Gi([0-9]+\/){2}[0-9]+'
	regex_description=r'[NRJPASEP]+[0-9]+[A-K0-9]+-[0-9]+'

	# Building DHCP Dictionnary

	for vlan in list(Vlans.keys()):
		f=open('../dhcpd-'+str(vlan)+'.conf')
		Content=f.read().split('}')
		for item in Content:
			matches=re.finditer(regex_MAC, item, re.MULTILINE)
			for matchNum, match in enumerate(matches, start=1):
				tmp=str(match.group())
				tmp_dict['mac']=tmp[:2]+tmp[3:5]+'.'+tmp[6:8]+tmp[9:11]+'.'+tmp[12:14]+tmp[15:17]
			matches=re.finditer(regex_IP, item, re.MULTILINE)
			for matchNum, match in enumerate(matches, start=1):
				matches2=re.finditer(regex_raw_ip, match.group(), re.MULTILINE)
				for mn, mat in enumerate(matches2, start=1):
					tmp_dict['ip']=str(mat.group())
			matches=re.finditer(regex_hostname, item, re.MULTILINE)
			for matchNum, match in enumerate(matches, start=1):
				tmp_dict['hostname']=str(match.group())
			tmp_dict['departement']=Vlans[vlan]
			tmp_dict['vlan']=vlan
			if tmp_dict != {}:
				Users.append(tmp_dict)
			tmp_dict={}
		DHCP_Dict[Vlans[vlan]]=Del_Duplicate(Users)[:-1]

	f=open('./Dictionnaire DHCP','w')
	for k,v in DHCP_Dict.items():
		print(k)
		f.write(str(k)+'\n')
		for item in v:
			print(item)
			f.write(str(item)+'\n')
	f.close()

def dhcp2ods(path):

	# Converting Raw dhcp dictionnary to proper Ods file


	# Reading raw file content to extract informations
	f=open(path,'r')

	# Regular Expressions definitions to treat faster as possible the dictionnary

	regex_mac=r'([0-9a-fA-F]{4}\.){2}[0-9a-fA-F]{4}'
	regex_ip=r'([0-9]+\.){3}[0-9]+'
	regex_hostname=r'(?<=hostname.....)[A-Za-z0-9-]*'
	regex_departement=r'(?<=departement.{4})[A-Za-z0-9-]*'
	regex_vlan=r'(?<=vlan.{3})[A-Za-z0-9-]*'

	# Variables Definition

	Mac_list=[]
	Ip_list=[]
	Host_list=[]
	Departement_list=[]
	Vlan_list=[]

	# Reading and filtering raw Content to dress Informations Lists
	Content=f.read()
	matches=re.finditer(regex_mac, Content, re.MULTILINE)
	for matchNum, match in enumerate(matches, start=1):
		Mac_list.append(match.group())
	matches=re.finditer(regex_ip, Content, re.MULTILINE)
	for matchNum, match in enumerate(matches, start=1):
		Ip_list.append(match.group())
	matches=re.finditer(regex_hostname, Content, re.MULTILINE)
	for matchNum, match in enumerate(matches, start=1):
		Host_list.append(match.group())	
	matches=re.finditer(regex_departement, Content, re.MULTILINE)
	for matchNum, match in enumerate(matches, start=1):
		Departement_list.append(match.group())	
	matches=re.finditer(regex_vlan, Content, re.MULTILINE)
	for matchNum, match in enumerate(matches, start=1):
		Vlan_list.append(match.group())	

	# Package lists as a excel file content to write
	to_write=[['Adresse mac','Adresse ip','Hostname','Dpt','Vlan id']]
	line=[]
	for i in range(len(Mac_list)):
		line=[str(Mac_list[i]),str(Ip_list[i]),str(Host_list[i]),str(Departement_list[i]),str(Vlan_list[i])]
		if not str(Departement_list[i])=='GUEST':
			to_write.append(line)
		line=[]
	Content={'Sheet 1':to_write}
	book = p.Book(Content)

	# Saving a pre-version of the dictionnary
	book.save_as('DHCP_dictionnary.xlsx')

	# Adjusting automatic size for columns
	Wb=op.load_workbook(filename='DHCP_dictionnary.xlsx')
	for Ws in Wb.worksheets:
		for col in Ws.columns:
			maxi=0
			column=op.utils.get_column_letter(col[0].column)
			for cell in col:
				try:
					if(len(str(cell.value)) > maxi):
						maxi=len(cell.value)
				except:
					pass 
			adj_width=(maxi + 2)*1.2
			if(column=='C'):
				adj_width=(maxi)
			elif(column=='D'):
				adj_width=maxi
			Ws.column_dimensions[column].width = adj_width
		Ws.showGridLines = True

	# Defining layout for the document
	border=op.styles.borders.Border(left=op.styles.borders.Side(style='medium'), 
                     right=op.styles.borders.Side(style='medium'), 
                     top=op.styles.borders.Side(style='medium'), 
                     bottom=op.styles.borders.Side(style='medium'))
	border3=op.styles.borders.Border(left=op.styles.borders.Side(style='thin'), 
	                     right=op.styles.borders.Side(style='thin'), 
	                     top=op.styles.borders.Side(style='thin'), 
	                     bottom=op.styles.borders.Side(style='thin'))
	font=op.styles.Font(color="00333333",size=12,bold=True)
	font2=op.styles.Font(color="00333333",size=11,bold=False)
	font3=op.styles.Font(color="00333300",italic=True)
	fill = op.styles.PatternFill("solid",fgColor="f8f8f8")
	fill2 = op.styles.PatternFill("solid",fgColor="f8f8f8")

	for i in range(1,6):
		Ws.cell(row=1,column=i).border=border
		Ws.cell(row=1,column=i).font=font
		Ws.cell(row=1,column=i).fill=fill

	for i in range(2,Ws.max_row+1):
		if(i<Ws.max_row):
			Ws.cell(row=i,column=1).fill=fill2
			Ws.cell(row=i,column=2).fill=fill2
			Ws.cell(row=i,column=3).fill=fill2
			Ws.cell(row=i,column=4).fill=fill2
			Ws.cell(row=i,column=5).fill=fill2

	# Saving file as xlsx file (temporary excel file)
	Wb.save(filename='DHCP_dictionnary.xlsx')

	# Convert to .ods file and delete temporary file
	os.system('soffice --headless --convert-to ods *.xlsx')
	os.system('rm *.xlsx')

Get_Users_Info()
dhcp2ods('./Dictionnaire DHCP')